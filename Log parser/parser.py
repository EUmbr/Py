import openpyxl
from openpyxl.styles import Font
import os
import io

files = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.log')]
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
err_font = Font(color='FF0000')
resp_err_font = Font(color='ab08fe')

row = 1
for f in files:
    cur_log = sheet.cell(row=row, column=1)
    cur_log.value = f
    cur_log.font = Font(bold=True)
    row += 1
    with io.open(f, encoding='utf-8') as log:
        for line in log:
            args = line.split('][')
            args[6] = args[6].split(' :: ')
            if args[6][0] == 'ERROR':
                sheet.cell(row=row, column=1).value = args[0][1:]
                sheet.cell(row=row, column=2).value = args[2]
                sheet.cell(row=row, column=3).value = args[5]
                err = sheet.cell(row=row, column=4)
                err.value = args[6][0]
                err.font = err_font
                sheet.cell(row=row, column=5).value = args[6][1][:-2]
                row += 1
            elif args[6][0] == 'RESPONSE' and args[6][3].split()[-1][0] != '2':
                sheet.cell(row=row, column=1).value = args[0][1:]
                sheet.cell(row=row, column=2).value = args[2]
                sheet.cell(row=row, column=3).value = args[5]
                err = sheet.cell(row=row, column=4)
                err.value = args[6][3]
                err.font = resp_err_font
                sheet.cell(row=row, column=5).value = args[6][1] + ' ' + args[6][2]
                sheet.cell(row=row, column=6).value = args[6][-1][:-2]
                row += 1
        if row == cur_log.row+1:
            sheet.cell(row=row, column=1).value = 'Нет ошибок'
        row += 3

sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 6
sheet.column_dimensions['D'].width = 24
sheet.column_dimensions['E'].width = 130

wb.save('logs.xlsx')
